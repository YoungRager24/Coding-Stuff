import openpyxl

#opening the workbook
wb = openpyxl.load_workbook("Furniture.xlsx")
sheet = wb["Furniture"]

#table types and their updated prices
price_updates = {"E":84.00, "D":184.00, "A":80.40}

#for loop goes in and updates prices based on created dictionary
for rowNum in range(4, sheet.max_row +1):
    tableType = sheet.cell(row=rowNum, column=5).value
    if tableType in price_updates:
        sheet.cell(row=rowNum, column=6).value = price_updates[tableType]

wb.save("updatedFurniture.xlsx")#saves new workbook with updated prices 
