import openpyxl

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

#produce types and their updated prices
price_updates = {"Garlic":5.50, "Celery":2.50, "Lemon":6.00}

for rowNum in range(2, sheet.max_row +1):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in price_updates:
        sheet.cell(row=rowNum, column=2).value = price_updates[produceName]

wb.save("updatedProduceSales.xlsx")
