import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb["Sheet1"]
print(sheet)

print(sheet["A1"])
print(sheet["A1"].value)
print(sheet["B1"].value)

title = sheet.title
print(title)

sheet.cell(row=1, column=2)
mydata = sheet.cell(row=1, column=2).value
print(mydata)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print("The max number of rows:", sheet.max_row)
print("The max number of columns:", sheet.max_column)

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(column_index_from_string("A"))

data = tuple(sheet["A1":"C3"])
print(data)

for rowOfCells in sheet["A1":"C3"]:
    for cellObject in rowOfCells:
        print(cellObject.coordinate, cellObject.value)
    print("....End of row....")
